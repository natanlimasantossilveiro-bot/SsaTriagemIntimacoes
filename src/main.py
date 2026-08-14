"""
SsaTriagemIntimacoes — organiza a planilha bruta de intimações do sistema
Publicações Online em uma planilha nova, mais limpa e priorizada.

Uso:
    python src/main.py input/NOME_DO_ARQUIVO.xlsx
"""

import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

# garante que este diretório (src/) esteja no sys.path, tanto rodando como
# script (`python src/main.py`) quanto importado como submódulo do pacote
# `src` (webapp.py, via `uvicorn src.webapp:app`).
sys.path.insert(0, str(Path(__file__).resolve().parent))

import config

MARCADOR_TEOR_INCOMPLETO = "Código dos itens com teor incompleto e/ou não exportado:"

# "Data Intimação: 2026-08-22 23:59:59.999 Prazo: 15 dias úteis" — usado como
# fallback quando Prazo inicial/Prazo final vêm vazios na planilha bruta.
PADRAO_PRAZO_TEXTO = re.compile(
    r"Data Intima[çc][ãa]o:\s*(\d{4}-\d{2}-\d{2})\s+[\d:.]+\s*Prazo:\s*(.+?)(?:\s{2,}|$)"
)


def carregar_planilha_bruta(caminho: Path):
    """
    Lê a aba de dados (Relatório) e a aba de metadados (Resumo).
    Retorna (df, codigos_incompletos).
    """
    df = pd.read_excel(caminho, sheet_name=config.ABA_DADOS, header=config.LINHA_CABECALHO - 1)
    df = df.dropna(subset=[config.COLUNA_PROCESSO]).reset_index(drop=True)

    resumo = pd.read_excel(caminho, sheet_name=config.ABA_RESUMO, header=None)

    codigos_incompletos = []
    primeira_coluna = resumo.iloc[:, 0].astype(str).str.strip()
    linhas_marcador = resumo.index[primeira_coluna == MARCADOR_TEOR_INCOMPLETO]
    if len(linhas_marcador) > 0:
        # a linha seguinte ao marcador é o cabeçalho da mini-tabela; os dados
        # começam duas linhas depois e vão até a primeira linha vazia.
        inicio_dados = linhas_marcador[0] + 2
        for i in range(inicio_dados, len(resumo)):
            valor = resumo.iloc[i, 0]
            if pd.isna(valor) or str(valor).strip() == "":
                break
            codigos_incompletos.append(str(valor).strip())

    return df, codigos_incompletos


def _remover_acentos(texto: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c))


def marcar_palavras_chave(df):
    """
    Busca (case-insensitive, sem acento) cada termo de config.PALAVRAS_CHAVE
    dentro de config.COLUNA_CONTEUDO. Funciona normalmente com a lista vazia.
    Adiciona "Palavras-chave encontradas" (nomes) e "Peso das palavras-chave"
    (soma dos pesos dos termos que bateram — usada como desempate em
    calcular_prioridade).
    """
    df = df.copy()
    termos = [
        (termo, _remover_acentos(termo).lower(), peso) for termo, peso in config.PALAVRAS_CHAVE
    ]

    def buscar(conteudo):
        if not isinstance(conteudo, str) or not conteudo or not termos:
            return "", 0
        texto_normalizado = _remover_acentos(conteudo).lower()
        encontrados = [(termo, peso) for termo, termo_norm, peso in termos if termo_norm in texto_normalizado]
        nomes = ", ".join(termo for termo, _ in encontrados)
        peso_total = sum(peso for _, peso in encontrados)
        return nomes, peso_total

    resultado = df[config.COLUNA_CONTEUDO].apply(buscar)
    df["Palavras-chave encontradas"] = resultado.apply(lambda r: r[0])
    df["Peso das palavras-chave"] = resultado.apply(lambda r: r[1])
    return df


def marcar_repetidos(df):
    """
    Marca processos repetidos (Qtde repetições (processo)) e conteúdo
    idêntico entre linhas de processos DIFERENTES (Conteúdo duplicado?).
    """
    df = df.copy()

    df["Qtde repetições (processo)"] = df.groupby(config.COLUNA_PROCESSO)[config.COLUNA_PROCESSO].transform(
        "count"
    )

    conteudo = df[config.COLUNA_CONTEUDO].fillna("").astype(str).str.strip()
    com_conteudo = conteudo != ""

    processos_por_conteudo = df.loc[com_conteudo].groupby(conteudo[com_conteudo])[
        config.COLUNA_PROCESSO
    ].nunique()
    conteudos_duplicados = set(processos_por_conteudo[processos_por_conteudo > 1].index)

    df["Conteúdo duplicado?"] = conteudo.apply(lambda c: "sim" if c in conteudos_duplicados else "não")
    return df


def marcar_teor_incompleto(df, codigos_incompletos: list[str]):
    """
    Marca Teor incompleto? = sim para os códigos vindos da aba Resumo.
    """
    df = df.copy()
    codigos_set = {str(c).strip() for c in codigos_incompletos}
    codigo_str = df[config.COLUNA_CODIGO].astype(str).str.strip()
    df["Teor incompleto?"] = codigo_str.apply(lambda c: "sim" if c in codigos_set else "não")
    return df


def _extrair_prazo_do_texto(conteudo):
    if not isinstance(conteudo, str):
        return None
    m = PADRAO_PRAZO_TEXTO.search(conteudo)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y-%m-%d")
    except ValueError:
        return None


def calcular_prioridade(df):
    """
    Ordena diretamente pela data-limite do prazo (coluna "Data limite
    (prazo)"), do prazo mais próximo de vencer para o mais distante — não
    pela Data de disponibilização/evento, que não reflete a duração real de
    cada prazo (duas linhas na mesma data de evento podem ter prazos de
    durações bem diferentes). O prazo continua sendo o critério principal;
    "Peso das palavras-chave" (ver marcar_palavras_chave) só desempata entre
    linhas com prazo igual (ex.: duas sem data-limite identificada).

    Além disso, agrupa as linhas do mesmo processo: o grupo aparece na
    posição da sua intimação mais urgente (menor data-limite / maior peso
    entre as do processo), e as demais do mesmo processo ficam logo em
    seguida, mesmo que individualmente fossem menos prioritárias.

    Quando Prazo inicial/final vêm vazios, a data-limite é extraída via
    regex do texto de Conteúdo (ver PADRAO_PRAZO_TEXTO) como fallback.
    "Origem do prazo" ("coluna", "texto" ou "não encontrado") registra de
    onde veio cada data. Linhas sem data-limite identificada vão para o
    final da lista.
    """
    df = df.copy()

    prazo_final_col = pd.to_datetime(df[config.COLUNA_PRAZO_FINAL], format="%d/%m/%Y", errors="coerce")

    origens = []
    datas_limite = []
    for idx in df.index:
        pf = prazo_final_col.loc[idx]
        if pd.notna(pf):
            origens.append("coluna")
            datas_limite.append(pf)
            continue
        data_texto = _extrair_prazo_do_texto(df.at[idx, config.COLUNA_CONTEUDO])
        if data_texto is not None:
            origens.append("texto")
            datas_limite.append(data_texto)
        else:
            origens.append("não encontrado")
            datas_limite.append(pd.NaT)

    df["Origem do prazo"] = origens
    df["Data limite (prazo)"] = pd.Series(datas_limite, index=df.index)

    hoje = pd.Timestamp(datetime.now().date())
    df["Prazo em aberto?"] = df["Data limite (prazo)"].apply(
        lambda d: "sim" if pd.notna(d) and d.normalize() >= hoje else "não"
    )

    # posição do grupo (mesmo Nº do processo): pela intimação mais urgente
    # dele — menor data-limite entre as linhas do processo e, em caso de
    # empate, o maior peso de palavra-chave entre elas.
    df["_grupo_data_limite"] = df.groupby(config.COLUNA_PROCESSO)["Data limite (prazo)"].transform("min")
    df["_grupo_peso"] = df.groupby(config.COLUNA_PROCESSO)["Peso das palavras-chave"].transform("max")

    df = df.sort_values(
        by=[
            "_grupo_data_limite",
            "_grupo_peso",
            config.COLUNA_PROCESSO,
            "Data limite (prazo)",
            "Peso das palavras-chave",
        ],
        ascending=[True, False, True, True, False],
        na_position="last",
    ).drop(columns=["_grupo_data_limite", "_grupo_peso"]).reset_index(drop=True)

    return df


CORES_SINALIZADORES = {
    "teor incompleto": "FFCDD2",
    "processo repetido": "BBDEFB",
    "conteúdo duplicado": "D1C4E9",
    "palavra-chave": "FFF9C4",
}


def _montar_sinalizadores(row):
    flags = []
    if row.get("Teor incompleto?") == "sim":
        flags.append("teor incompleto")
    if row.get("Qtde repetições (processo)", 1) > 1:
        flags.append("processo repetido")
    if row.get("Conteúdo duplicado?") == "sim":
        flags.append("conteúdo duplicado")
    if row.get("Palavras-chave encontradas"):
        flags.append("palavra-chave")
    return ", ".join(flags)


def gerar_planilha_final(df, caminho_saida: Path):
    """
    Exporta para output/ com fonte Arial, cabeçalho congelado, autofiltro e
    preenchimento condicional. Nunca sobrescreve o arquivo de entrada (o
    caminho de saída sempre fica em output/, isolado de input/). Espera que
    df já tenha a coluna "Sinalizadores" (ver processar_planilha).
    """
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Intimações")

    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active

    fonte = Font(name="Arial", size=10)
    fonte_cabecalho = Font(name="Arial", size=10, bold=True)
    for linha in ws.iter_rows():
        for celula in linha:
            celula.font = fonte
    for celula in ws[1]:
        celula.font = fonte_cabecalho

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    colunas = {celula.value: celula.column for celula in ws[1]}
    total_colunas = ws.max_column

    # ordem de precedência: se a linha acumula mais de um sinalizador, a cor
    # mostrada segue essa ordem (a coluna "Sinalizadores" mantém a lista completa).
    for i in range(len(df)):
        linha_excel = i + 2
        row = df.iloc[i]
        if row.get("Teor incompleto?") == "sim":
            cor = CORES_SINALIZADORES["teor incompleto"]
        elif row.get("Qtde repetições (processo)", 1) > 1:
            cor = CORES_SINALIZADORES["processo repetido"]
        elif row.get("Conteúdo duplicado?") == "sim":
            cor = CORES_SINALIZADORES["conteúdo duplicado"]
        elif row.get("Palavras-chave encontradas"):
            cor = CORES_SINALIZADORES["palavra-chave"]
        else:
            continue

        preenchimento = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        for col in range(1, total_colunas + 1):
            ws.cell(row=linha_excel, column=col).fill = preenchimento

    if config.COLUNA_CONTEUDO in colunas:
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(colunas[config.COLUNA_CONTEUDO])
        ].width = 60
    if "Sinalizadores" in colunas:
        ws.column_dimensions[openpyxl.utils.get_column_letter(colunas["Sinalizadores"])].width = 30

    wb.save(caminho_saida)


@dataclass
class ResultadoProcessamento:
    caminho_saida: Path
    total_linhas: int
    processos_repetidos: int
    teor_incompleto: int
    df: pd.DataFrame


def processar_planilha(caminho_entrada: Path, pasta_saida: Path = config.PASTA_OUTPUT) -> ResultadoProcessamento:
    """
    Pipeline completo: carrega a planilha bruta, aplica as marcações e a
    priorização, e gera a planilha final em pasta_saida. Reaproveitado pelo
    CLI (main()) e pelo webapp (rota POST /processar, que também usa o
    DataFrame retornado pra alimentar a tela de intimações via intimacoes_db).
    """
    nome_saida = f"{caminho_entrada.stem}_organizado.xlsx"
    caminho_saida = pasta_saida / nome_saida

    df, codigos_incompletos = carregar_planilha_bruta(caminho_entrada)
    df = marcar_palavras_chave(df)
    df = marcar_repetidos(df)
    df = marcar_teor_incompleto(df, codigos_incompletos)
    df = calcular_prioridade(df)
    df["Sinalizadores"] = df.apply(_montar_sinalizadores, axis=1)
    gerar_planilha_final(df, caminho_saida)

    return ResultadoProcessamento(
        caminho_saida=caminho_saida,
        total_linhas=len(df),
        df=df,
        processos_repetidos=int(df[df["Qtde repetições (processo)"] > 1][config.COLUNA_PROCESSO].nunique()),
        teor_incompleto=int((df["Teor incompleto?"] == "sim").sum()),
    )


def main():
    if len(sys.argv) != 2:
        print("Uso: python src/main.py input/NOME_DO_ARQUIVO.xlsx")
        sys.exit(1)

    caminho_entrada = Path(sys.argv[1])
    resultado = processar_planilha(caminho_entrada)

    print(f"Planilha organizada gerada em: {resultado.caminho_saida}")


if __name__ == "__main__":
    main()